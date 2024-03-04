from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math
import json

def DamageCalulate(power, currentAtk, currentDef, stab, resist):
    return math.floor(0.5*power*currentAtk/currentDef*stab*resist)+1

def CPCalulate(baseAtk, baseDef, baseSta, CPMTable):
    bestIV = [0,0,0,0,0]
    maxstatproduct = 0
    for AtkIV in range(0,16):
        for DefIV in range(0,16):
            for StaIV in range(0,16):
                CP = 0
                lv = 1
                for i in range(20,101):
                    CP = math.floor(max(10,((baseAtk + AtkIV)*((baseDef + DefIV)**0.5)*((baseSta + StaIV)**0.5)*(CPMTable[i][1])**2)/10))
                    if CP > 1500:
                        break
                    lv = CPMTable[i][0]
                    statproduct = round((baseAtk + AtkIV)*CPMTable[i][1],2)*round((baseDef + DefIV)*CPMTable[i][1],2)*math.floor((baseSta + StaIV)*CPMTable[i][1])
                    if statproduct >= maxstatproduct:
                        maxstatproduct = statproduct
                        bestIV = [AtkIV, DefIV, StaIV, lv, CP, i, statproduct]
    return bestIV 



STAB = 1.2

Neutral = 1
SuperEffective = 1.6
Not_Very_Effective = 0.625
Immune = 0.390625


with open('CPMTable.json', 'r') as json_file:
    CPMTable = json.load(json_file)
print(CPMTable[100][0])

#開啟要存的excel檔
pokemonstat = load_workbook('種族值列表GO02.xlsx')
statsSheet = pokemonstat.get_sheet_by_name('Pokemon_Stats')
#種族值列表
pokemonstats = pokemonstat.active

#for attacker in range(2,statsSheet.max_row+2):
#    atkpoke = statsSheet['B'+str(attacker)].value 
 #   [atkpokeAtkIV, atkpokeDefIV, atkpokeStaIV, atkpokelv, atkpokeCP, atkpokeCPMidx] = CPCalulate(statsSheet['N'+str(attacker)].value, statsSheet['Q'+str(attacker)].value, statsSheet['S'+str(attacker)].value, CPMTable)
 #   print([atkpokeAtkIV, atkpokeDefIV, atkpokeStaIV, atkpokelv, atkpokeCP, atkpokeCPMidx])
 #   for defender in range(2,statsSheet.max_row+2):
 #       defpoke = statsSheet['B'+str(defender)].value 
 #       [defpokeAtkIV, defpokeDefIV, defpokeStaIV, defpokelv, defpokeCP, defpokeCPMidx] = CPCalulate(statsSheet['N'+str(attacker)].value, statsSheet['Q'+str(attacker)].value, statsSheet['S'+str(attacker)].value, CPMTable)
#
 #       atkpokecurrentAtk = (statsSheet['N'+str(attacker)].value + atkpokeAtkIV)*CPMTable[atkpokeCPMidx][1]
 #       atkpokecurrentDef = (statsSheet['Q'+str(attacker)].value + atkpokeDefIV)*CPMTable[atkpokeCPMidx][1]
 #       atkpokecurrentSta = (statsSheet['S'+str(attacker)].value + atkpokeStaIV)*CPMTable[atkpokeCPMidx][1]

 #       defpokecurrentAtk = (statsSheet['N'+str(defender)].value + defpokeAtkIV)*CPMTable[defpokeCPMidx][1]
 #       defpokecurrentDef = (statsSheet['Q'+str(defender)].value + defpokeDefIV)*CPMTable[defpokeCPMidx][1]
 #       defpokecurrentSta = (statsSheet['S'+str(defender)].value + defpokeStaIV)*CPMTable[defpokeCPMidx][1]

 #       DamageCalulate(power, currentAtk, currentDef, stab, resist)
    

#Default PVPIv
for attacker in range(2,statsSheet.max_row):
    atkpoke = statsSheet['B'+str(attacker)].value 
    [atkpokeAtkIV, atkpokeDefIV, atkpokeStaIV, atkpokelv, atkpokeCP, atkpokeCPMidx, atkpokestatproduct] = CPCalulate(statsSheet['N'+str(attacker)].value, statsSheet['Q'+str(attacker)].value, statsSheet['S'+str(attacker)].value, CPMTable)
    print([atkpokeAtkIV, atkpokeDefIV, atkpokeStaIV, atkpokelv, atkpokeCP, atkpokeCPMidx])
    with open('Pokemon_1500_default.json', 'r', encoding="utf-8") as json_file:
        all_data = json.load(json_file)

    print(atkpoke)
    all_data[f'{atkpoke}'] = {}
    all_data[f'{atkpoke}']['index'] = statsSheet['A'+str(attacker)].value
    all_data[f'{atkpoke}']['DefaultAtkIV'] = atkpokeAtkIV
    all_data[f'{atkpoke}']['DefaulDefIV'] = atkpokeDefIV
    all_data[f'{atkpoke}']['DefaulStaIV'] = atkpokeStaIV
    all_data[f'{atkpoke}']['DefaulLV'] = atkpokelv
    all_data[f'{atkpoke}']['DefaulCP'] = atkpokeCP
    all_data[f'{atkpoke}']['CurrentAtk'] = (statsSheet['N'+str(attacker)].value + atkpokeAtkIV)*CPMTable[atkpokeCPMidx][1]
    all_data[f'{atkpoke}']['CurrentDef'] = (statsSheet['Q'+str(attacker)].value + atkpokeDefIV)*CPMTable[atkpokeCPMidx][1]
    all_data[f'{atkpoke}']['CurrentHP'] = (statsSheet['S'+str(attacker)].value + atkpokeStaIV)*CPMTable[atkpokeCPMidx][1]
    

    with open('pokemon.json', 'r', encoding="utf-8") as p:
        pok = json.load(p)
    
    for i in range(0,len(pok)):
        if pok[i]['dex'] == statsSheet['A'+str(attacker)].value:
            all_data[f'{atkpoke}']['type'] = pok[i]['types']
            all_data[f'{atkpoke}']['fastMoves'] = pok[i]['fastMoves']
            all_data[f'{atkpoke}']['chargedMoves'] = pok[i]['chargedMoves']
            if 'tags' in pok[i]:
                if 'shadoweligible' in pok[i]['tags']:
                    all_data[f'{atkpoke}']['shadow'] = True
                elif 'shadow' in pok[i]['tags']:
                    all_data[f'{atkpoke}']['shadow'] = True
                else:
                    all_data[f'{atkpoke}']['shadow'] = False

            break

    with open('Pokemon_1500_default.json', 'w', encoding="utf-8") as json_file:
        json.dump(all_data, json_file, ensure_ascii=False)
print('finish')

#for row in range(2,913):  
#    for i in range(2,913):
#        #print(lead['A'+str(i)].value)
#        if dataset['A'+str(row)].value == lead['A'+str(i)].value:     
#           dataset['C'+str(row)].value = lead['B'+str(i)].value






