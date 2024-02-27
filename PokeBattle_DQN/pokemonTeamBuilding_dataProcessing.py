from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


#開啟要存的excel檔
wb1 = load_workbook('Pokemon_Team_Building_dataset.xlsx')
wb1s = wb1.active
print(wb1.worksheets[2])
dataset = wb1.worksheets[1]
lead = wb1.worksheets[2]
switch = wb1.worksheets[3]
closer = wb1.worksheets[4]


for row in range(2,913):  
    for i in range(2,913):
        #print(lead['A'+str(i)].value)
        if dataset['A'+str(row)].value == lead['A'+str(i)].value:     
            dataset['C'+str(row)].value = lead['B'+str(i)].value

for row in range(2,913):  
    for i in range(2,913):
        #print(lead['A'+str(i)].value)
        if dataset['A'+str(row)].value == switch['A'+str(i)].value:     
            dataset['D'+str(row)].value = switch['B'+str(i)].value

for row in range(2,913):
    for i in range(2,913):
        #print(lead['A'+str(i)].value)
        if dataset['A'+str(row)].value == closer['A'+str(i)].value:     
            dataset['E'+str(row)].value = closer['B'+str(i)].value
#存檔
wb1.save('Pokemon_Team_Building_dataset.xlsx')
print("finish")

